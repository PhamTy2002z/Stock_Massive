"""Bounded PDF, XLSX, and DOCX extraction with exact source locators."""

from __future__ import annotations

import hashlib
import io
import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import PurePath
from typing import Any
from xml.etree import ElementTree

from .contracts import (
    EvidenceKind,
    EvidenceLocation,
    EvidenceRef,
    SourceClass,
    build_evidence_ref,
)

PDF_MEDIA_TYPE = "application/pdf"
XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
DOCX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
)
SUPPORTED_DOCUMENT_TYPES = frozenset(
    {PDF_MEDIA_TYPE, XLSX_MEDIA_TYPE, DOCX_MEDIA_TYPE}
)


@dataclass(frozen=True)
class DocumentQuotas:
    max_input_bytes: int = 8 * 1024 * 1024
    max_output_chars: int = 200_000
    max_excerpt_chars: int = 8_000
    max_pdf_pages: int = 80
    max_sheets: int = 32
    max_rows_per_sheet: int = 2_000
    max_columns_per_sheet: int = 256
    max_cells_per_sheet: int = 20_000
    max_zip_members: int = 2_000
    max_zip_uncompressed_bytes: int = 40 * 1024 * 1024

    def __post_init__(self) -> None:
        for name, value in self.__dict__.items():
            if value < 1:
                raise ValueError(f"{name} must be positive")


DEFAULT_DOCUMENT_QUOTAS = DocumentQuotas()


class DocumentParseError(ValueError):
    """A document is unsupported, invalid, encrypted, or exceeds a quota."""

    def __init__(self, code: str, detail: str) -> None:
        super().__init__(f"{code}: {detail}")
        self.code = code
        self.detail = detail


@dataclass(frozen=True)
class ParsedDocument:
    filename: str
    media_type: str
    content_sha256: str
    evidence: tuple[EvidenceRef, ...]
    truncated: bool
    warnings: tuple[str, ...]

    @property
    def extracted_chars(self) -> int:
        return sum(len(item.excerpt) for item in self.evidence)

    def to_payload(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "mediaType": self.media_type,
            "contentSha256": self.content_sha256,
            "evidence": [item.to_payload() for item in self.evidence],
            "truncated": self.truncated,
            "warnings": list(self.warnings),
        }


class _Collector:
    def __init__(
        self,
        *,
        filename: str,
        digest: str,
        observed_at: datetime,
        quotas: DocumentQuotas,
    ) -> None:
        self.filename = filename
        self.digest = digest
        self.observed_at = observed_at
        self.quotas = quotas
        self.evidence: list[EvidenceRef] = []
        self.output_chars = 0
        self.truncated = False
        self.warnings: list[str] = []

    def add(
        self,
        *,
        kind: EvidenceKind,
        location: EvidenceLocation,
        excerpt: str,
        label: str,
    ) -> bool:
        normalized = "\n".join(
            line.rstrip() for line in excerpt.replace("\x00", "").splitlines()
        ).strip()
        if not normalized:
            return True
        remaining = self.quotas.max_output_chars - self.output_chars
        if remaining <= 0:
            self.mark_truncated("document output reached the character quota")
            return False
        limit = min(remaining, self.quotas.max_excerpt_chars)
        if len(normalized) > limit:
            normalized = normalized[:limit].rstrip()
            self.mark_truncated(f"{label} was shortened to the excerpt quota")
        if not normalized:
            return False
        self.evidence.append(
            build_evidence_ref(
                kind=kind,
                source_class=SourceClass.USER_DOCUMENT,
                title=f"{self.filename} — {label}",
                source=self.filename,
                excerpt=normalized,
                content_sha256=self.digest,
                location=location,
                observed_at=self.observed_at,
            )
        )
        self.output_chars += len(normalized)
        return self.output_chars < self.quotas.max_output_chars

    def mark_truncated(self, warning: str) -> None:
        self.truncated = True
        if warning not in self.warnings:
            self.warnings.append(warning)

    def finish(self, media_type: str) -> ParsedDocument:
        if not self.evidence:
            raise DocumentParseError(
                "no_extractable_text", "the document contains no extractable text"
            )
        return ParsedDocument(
            filename=self.filename,
            media_type=media_type,
            content_sha256=self.digest,
            evidence=tuple(self.evidence),
            truncated=self.truncated,
            warnings=tuple(self.warnings),
        )


def _safe_filename(filename: str) -> str:
    name = PurePath(filename.replace("\\", "/")).name.strip()
    if not name or name in {".", ".."}:
        raise DocumentParseError("invalid_filename", "filename cannot be blank")
    return name


def _check_input(content: bytes, quotas: DocumentQuotas) -> None:
    if not content:
        raise DocumentParseError("empty_document", "document content is empty")
    if len(content) > quotas.max_input_bytes:
        raise DocumentParseError(
            "input_too_large",
            f"document exceeds {quotas.max_input_bytes} bytes",
        )


def _check_zip(content: bytes, quotas: DocumentQuotas) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > quotas.max_zip_members:
                raise DocumentParseError(
                    "archive_too_large",
                    f"archive has more than {quotas.max_zip_members} members",
                )
            uncompressed = sum(member.file_size for member in members)
            if uncompressed > quotas.max_zip_uncompressed_bytes:
                raise DocumentParseError(
                    "archive_too_large",
                    "archive exceeds the uncompressed byte quota",
                )
    except zipfile.BadZipFile as exc:
        raise DocumentParseError("invalid_archive", "document is not a valid ZIP") from exc


def _parse_pdf(content: bytes, collector: _Collector) -> ParsedDocument:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - deployment packaging failure
        raise DocumentParseError(
            "dependency_missing", "PDF extraction requires pypdf"
        ) from exc

    try:
        reader = PdfReader(io.BytesIO(content), strict=True)
        if reader.is_encrypted:
            raise DocumentParseError(
                "encrypted_document", "encrypted PDFs are not supported"
            )
        if len(reader.pages) > collector.quotas.max_pdf_pages:
            collector.mark_truncated(
                f"only the first {collector.quotas.max_pdf_pages} PDF pages were read"
            )
        for page_number, page in enumerate(
            reader.pages[: collector.quotas.max_pdf_pages], start=1
        ):
            try:
                text = page.extract_text() or ""
            except Exception as exc:  # one malformed page must not erase prior pages
                collector.warnings.append(
                    f"page {page_number} could not be extracted: {type(exc).__name__}"
                )
                continue
            if not collector.add(
                kind=EvidenceKind.DOCUMENT_PAGE,
                location=EvidenceLocation(page=page_number),
                excerpt=text,
                label=f"page {page_number}",
            ):
                break
    except DocumentParseError:
        raise
    except Exception as exc:
        raise DocumentParseError("invalid_pdf", "PDF could not be parsed") from exc
    return collector.finish(PDF_MEDIA_TYPE)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def _parse_xlsx(content: bytes, collector: _Collector) -> ParsedDocument:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - deployment packaging failure
        raise DocumentParseError(
            "dependency_missing", "XLSX extraction requires openpyxl"
        ) from exc

    _check_zip(content, collector.quotas)
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise DocumentParseError("invalid_xlsx", "workbook could not be parsed") from exc

    try:
        sheets = workbook.worksheets
        if len(sheets) > collector.quotas.max_sheets:
            collector.mark_truncated(
                f"only the first {collector.quotas.max_sheets} sheets were read"
            )
        for sheet in sheets[: collector.quotas.max_sheets]:
            cell_count = 0
            row_count = 0
            max_row = min(sheet.max_row, collector.quotas.max_rows_per_sheet)
            max_column = min(
                sheet.max_column, collector.quotas.max_columns_per_sheet
            )
            if sheet.max_row > collector.quotas.max_rows_per_sheet:
                collector.mark_truncated(
                    f"sheet {sheet.title!r} reached the row quota"
                )
            if sheet.max_column > collector.quotas.max_columns_per_sheet:
                collector.mark_truncated(
                    f"sheet {sheet.title!r} reached the column quota"
                )
            reached_cell_quota = False
            for row_number, row in enumerate(
                sheet.iter_rows(max_row=max_row, max_col=max_column), start=1
            ):
                populated: list[tuple[int, str]] = []
                for column_number, cell in enumerate(row, start=1):
                    cell_count += 1
                    if cell_count > collector.quotas.max_cells_per_sheet:
                        collector.mark_truncated(
                            f"sheet {sheet.title!r} reached the cell quota"
                        )
                        reached_cell_quota = True
                        break
                    text = _cell_text(cell.value).strip()
                    if not text:
                        continue
                    populated.append((column_number, text))
                if reached_cell_quota:
                    break
                if not populated:
                    continue
                row_count += 1
                first_column = get_column_letter(populated[0][0])
                last_column = get_column_letter(populated[-1][0])
                cell_range = (
                    f"{first_column}{row_number}"
                    if first_column == last_column
                    else f"{first_column}{row_number}:{last_column}{row_number}"
                )
                excerpt = " | ".join(
                    f"{get_column_letter(column)}{row_number}={text}"
                    for column, text in populated
                )
                if not collector.add(
                    kind=EvidenceKind.DOCUMENT_CELL_RANGE,
                    location=EvidenceLocation(
                        sheet=sheet.title, cell_range=cell_range
                    ),
                    excerpt=excerpt,
                    label=f"{sheet.title}!{cell_range}",
                ):
                    break
            if not row_count:
                collector.warnings.append(f"sheet {sheet.title!r} was empty")
            if collector.output_chars >= collector.quotas.max_output_chars:
                break
    finally:
        workbook.close()
    return collector.finish(XLSX_MEDIA_TYPE)


_W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _xml_text(element: ElementTree.Element) -> str:
    return "".join(node.text or "" for node in element.iter(f"{_W}t")).strip()


def _paragraph_style(element: ElementTree.Element) -> str | None:
    style = element.find(f"./{_W}pPr/{_W}pStyle")
    if style is None:
        return None
    return style.attrib.get(f"{_W}val")


def _is_heading(style: str | None) -> bool:
    if style is None:
        return False
    compact = style.casefold().replace(" ", "")
    return compact.startswith("heading") or compact.startswith("tieude")


def _parse_docx(content: bytes, collector: _Collector) -> ParsedDocument:
    _check_zip(content, collector.quotas)
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            xml = archive.read("word/document.xml")
    except KeyError as exc:
        raise DocumentParseError(
            "invalid_docx", "word/document.xml is missing"
        ) from exc
    except zipfile.BadZipFile as exc:
        raise DocumentParseError("invalid_docx", "DOCX could not be parsed") from exc

    try:
        root = ElementTree.fromstring(xml)
    except ElementTree.ParseError as exc:
        raise DocumentParseError("invalid_docx", "document XML is invalid") from exc
    body = root.find(f"{_W}body")
    if body is None:
        raise DocumentParseError("invalid_docx", "document body is missing")

    section = "Document"
    block = 0
    table_number = 0
    for child in body:
        if child.tag == f"{_W}p":
            text = _xml_text(child)
            if not text:
                continue
            if _is_heading(_paragraph_style(child)):
                section = text
                continue
            block += 1
            if not collector.add(
                kind=EvidenceKind.DOCUMENT_SECTION,
                location=EvidenceLocation(section=section, block=block),
                excerpt=text,
                label=f"{section}, block {block}",
            ):
                break
        elif child.tag == f"{_W}tbl":
            table_number += 1
            for row_number, row in enumerate(child.findall(f"./{_W}tr"), start=1):
                values = [_xml_text(cell) for cell in row.findall(f"./{_W}tc")]
                if not any(values):
                    continue
                block += 1
                last_column = max(1, len(values))
                if not collector.add(
                    kind=EvidenceKind.DOCUMENT_CELL_RANGE,
                    location=EvidenceLocation(
                        sheet=f"Table {table_number}",
                        cell_range=f"R{row_number}C1:R{row_number}C{last_column}",
                        section=section,
                        block=block,
                    ),
                    excerpt=" | ".join(
                        f"C{column}={value}"
                        for column, value in enumerate(values, start=1)
                        if value
                    ),
                    label=f"{section}, table {table_number}, row {row_number}",
                ):
                    break
        if collector.output_chars >= collector.quotas.max_output_chars:
            break
    return collector.finish(DOCX_MEDIA_TYPE)


def parse_document(
    *,
    content: bytes,
    media_type: str,
    filename: str,
    quotas: DocumentQuotas = DEFAULT_DOCUMENT_QUOTAS,
    observed_at: datetime | None = None,
) -> ParsedDocument:
    """Extract a supported document without network, OCR, or unbounded reads."""

    _check_input(content, quotas)
    if media_type not in SUPPORTED_DOCUMENT_TYPES:
        raise DocumentParseError(
            "unsupported_media_type", f"unsupported media type: {media_type}"
        )
    when = observed_at or datetime.now(UTC)
    if when.utcoffset() is None:
        raise ValueError("observed_at must include a timezone")
    collector = _Collector(
        filename=_safe_filename(filename),
        digest=hashlib.sha256(content).hexdigest(),
        observed_at=when,
        quotas=quotas,
    )
    if media_type == PDF_MEDIA_TYPE:
        return _parse_pdf(content, collector)
    if media_type == XLSX_MEDIA_TYPE:
        return _parse_xlsx(content, collector)
    return _parse_docx(content, collector)
